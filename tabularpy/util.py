import copy
import html
import locale
import re
from collections import OrderedDict, defaultdict
from datetime import datetime, timedelta
from decimal import Decimal
from html.parser import HTMLParser
from pathlib import Path


def min_aggr(old, new, _):
    if new < old:
        return new
    else:
        return old


def max_aggr(old, new, _):
    if new > old:
        return new
    else:
        return old


def product_aggr(old, new, _):
    return old * new


def count_aggr(*args):
    return args[2]


def avg_aggr(mean, new, x):
    return mean + ((new - mean) / x)


def sum_aggr(old, new, _):
    return old + new


def clean_value(value, type_desc, settings):
    if value and settings.clean_values:
        value = str(value)
        type_desc = str(type_desc).lower()
        if value == "-" and ("str" not in type_desc or "varchar" not in type_desc):
            return None
        if (
            type_desc == "integer"
            or type_desc == "int"
            or type_desc == "bigint"
            or type_desc == "seconds"
        ):
            return int(value.replace(",", ""))
        elif type_desc == "float":
            return float(value.replace(",", ""))
        elif type_desc == "percent":
            value = Decimal(value.replace(",", "").replace("%", ""))
            if settings.divide_percent:
                value = value / 100
            return value
        elif type_desc == "money":
            return Decimal(value.replace(",", "").replace("$", ""))
        elif type_desc == "decimal" or type_desc == "numeric":
            return Decimal(value)
        elif type_desc == "bool":
            if isinstance(value, str):
                if value.lower() == "false":
                    value = False
                elif value.lower() == "true":
                    value = True
            return bool(value)
        elif type_desc == "date":
            return parse_date_time_string(value, settings.date_format)
        elif type_desc == "time":
            return parse_date_time_string(value, settings.time_format)
        elif type_desc == "interval":
            return parse_date_time_string(value)
        elif type_desc == "timestamp":
            return parse_date_time_string(value, settings.datetime_format)
    elif value == "":
        if settings.empty_string_is_none:
            return None
    return value


def format_value(value, type_desc, str_format=None):
    type_desc = str(type_desc).lower()
    if value is not None:
        if (
            type_desc == "integer"
            or type_desc == "int"
            or type_desc == "bigint"
            or type_desc == "seconds"
        ):
            return locale.format_string("%d", value, grouping=True)
        elif type_desc == "float" or type_desc == "decimal" or type_desc == "numeric":
            return (
                locale.format_string(
                    "%.4f", Decimal(str(value).replace(",", "")), grouping=True
                )
                .rstrip("0")
                .rstrip(".")
            )
        elif type_desc == "percent":
            value = value * 100
            s = str(value)
            v = (
                locale.format_string(
                    "%g", Decimal(s.rstrip("0").rstrip(".")), grouping=True
                )
                if "." in s
                else locale.format_string("%g", Decimal(s), grouping=True)
            )
            return f"{v}%"
        elif type_desc == "money":
            if value < 0:
                result = f"-{locale.currency(abs(value), grouping=True)}"
            else:
                result = locale.currency(value, grouping=True)
            return result
        elif type_desc == "date":
            if "month" in type_desc:
                return value.strftime("%m/%Y")
            elif "quarter" in type_desc:
                return datetime_to_quarter(value)
            else:
                return value.strftime("%m/%d/%Y")
        elif type_desc == "timestamp" or type_desc == "time" or type_desc == "interval":
            return value.strftime(str_format)
    return str(value)


def cast(value):
    if isinstance(value, float):
        return "float"
    elif isinstance(value, Decimal):
        return "decimal"
    elif isinstance(value, str):
        return "str"


def get_sql_query_types(query):
    t = OrderedDict()
    for column in query.column_descriptions:
        t[column["name"]] = column["type"]
    return t


def seconds_since_epoch(date_time):
    return (date_time - datetime(1970, 1, 1)).total_seconds()


def getitem(obj, index, default=None):
    try:
        value = obj[index]
        return value if value is not None else default
    except IndexError:
        return default


def getindex(obj, item, default=0):
    try:
        return obj.index(item)
    except ValueError:
        return default


def getindexes(obj, item):
    return [index for index, value in enumerate(obj) if value == item]


def rotate_clockwise(matrix, degree=90):
    if degree not in [0, 90, 180, 270, 360]:
        return
    return (
        matrix
        if not degree
        else rotate_clockwise(zip(*list(matrix)[::-1]), degree - 90)
    )


def horizontal_matrix_flip(matrix):
    return [list(reversed(row)) for row in matrix]


def transpose(matrix):
    return horizontal_matrix_flip(rotate_clockwise(matrix))


def find_duplicates(lst):
    temp = defaultdict(list)
    dups = {}
    for i, e in enumerate(lst):
        temp[e].append(i)
    for k, v in temp.items():
        if len(v) >= 2:
            dups[k] = v
    return dups


def merge_sort(array, index, key=None, reverse=False):
    size = len(array)
    if size < 2:
        return array
    if not key:
        key = lambda x: x
    middle = size // 2
    left = merge_sort(array[:middle], index, key, reverse)
    right = merge_sort(array[middle:], index, key, reverse)
    left.reverse()
    right.reverse()
    result = []
    while left and right:
        if not reverse:
            if key(left[-1][index]) <= key(right[-1][index]):
                result.append(left.pop())
            else:
                result.append(right.pop())
        else:
            if key(left[-1][index]) >= key(right[-1][index]):
                result.append(left.pop())
            else:
                result.append(right.pop())
    left.reverse()
    right.reverse()
    if left:
        result.extend(left)
    if right:
        result.extend(right)
    return result


def datetime_to_quarter(date_time):
    return "Q%s %s" % ((date_time.month - 1) // 3 + 1, date_time.year)


def parse_time_delta(s):
    if isinstance(s, str):
        if s == "0":
            return timedelta()
        if s is None:
            return None
        d = re.search(
            r"((?P<days>\d+) )?(?P<hours>\d+):(?P<minutes>\d+):(?P<seconds>\d+).(?P<deciseconds>\d+)",
            s,
        )
        if d is None:
            print(s, "could not be parsed")
            return None
        d = d.groupdict(0)
        seconds = 0.0
        if "days" in d:
            seconds += int(d["days"]) * 86400
        if "hours" in d:
            seconds += int(d["hours"]) * 3600
        if "minutes" in d:
            seconds += int(d["minutes"]) * 60
        if "seconds" in d:
            seconds += int(d["seconds"])
        if "deciseconds" in d:
            seconds += float(f".{d['deciseconds']}")
        return timedelta(seconds=seconds)
    return s


def parse_date_time_string(value, str_format=None):
    if isinstance(value, str):
        try:
            # noinspection PyUnresolvedReferences
            from dateutil.parser import parse

            value = parse(value)
        except ImportError:
            value = datetime.strptime(value, str_format)
        except ValueError:
            pass
    return value


class HtmlParser(HTMLParser):
    __slots__ = [
        "html",
        "_open_tags",
        "_open_tag_attrs",
        "_relevant_tags",
        "_row_num",
        "_col_num",
        "_data",
        "_hidden_rows",
        "_td",
    ]

    def __init__(self):
        super().__init__()
        self._open_tags = []
        self._open_tag_attrs = []
        self._relevant_tags = ("table", "thead", "tbody", "tfoot", "tr", "th", "td")
        self._row_num = -1
        self._col_num = -1
        self._data = []
        self._hidden_rows = []
        self._td = {"table": {}, "headers": [], "footers": [], "column_types": {}}

    def parse(self, htm):
        self.reset()
        self.feed(htm)
        return self._td

    def handle_starttag(self, tag, attrs):
        if tag in self._relevant_tags:
            self._open_tags.append(tag)
            self._open_tag_attrs.append({key: value for key, value in attrs})
            if "tbody" in self._open_tags:
                if tag == "tr":
                    self._row_num += 1
                    self._col_num = -1
                elif tag == "td":
                    self._col_num += 1

    def handle_endtag(self, tag):
        if tag in self._relevant_tags:
            last_open_tag = self._open_tags[-1]
            last_open_attrs = self._open_tag_attrs[-1]
            if "tr" == last_open_tag:
                for attr in last_open_attrs.values():
                    if "display: none" in attr:
                        self._hidden_rows.append(self._row_num)
            if "thead" in self._open_tags:
                if "th" == last_open_tag:
                    self._data = "".join(self._data)
                    if self._data == "":
                        self._data = str(self._col_num)
                    self._td["headers"].append(self._data)
                    self._td["table"][self._data] = []
                    self._data = []
            elif "tfoot" in self._open_tags:
                if "th" == last_open_tag:
                    self._data = "".join(self._data)
                    self._td["footers"].append(self._data)
                    self._data = []
            elif "tbody" in self._open_tags:
                if "td" == last_open_tag:
                    header = self._td["headers"][self._col_num]
                    self._data = "".join(self._data)
                    self._td["table"][header].append(self._data)
                    if header not in self._td["column_types"]:
                        self._td["column_types"][header] = last_open_attrs
                    self._data = []
            if tag == last_open_tag:
                self._open_tags.pop()
                self._open_tag_attrs.pop()

    def handle_data(self, data):
        self._data.append(data.strip())

    def handle_charref(self, ref):
        self.handle_entityref("#" + ref)

    def handle_entityref(self, ref):
        self.handle_data(html.unescape(f"&{ref};"))

    def error(self, message):
        print(message)


class BeautifulSoupParser(object):
    def __init__(self, parser="html.parser"):
        self.soup = None
        self.parser = parser

    def parse(self, htm):
        try:
            # noinspection PyUnresolvedReferences
            from bs4 import BeautifulSoup
        except ImportError:
            print("Need to install BeautifulSoup4 or use the standard HTML Parser")
            raise
        td = {"table": {}, "headers": [], "footers": [], "column_types": {}}
        self.soup = BeautifulSoup(htm, self.parser)
        self.normalize_html()
        table = self.soup.find("table")
        for x, header in enumerate(table.find("thead").find("tr").find_all("th")):
            if header.text == "":
                td["headers"].append(str(x))
                td["table"][str(x)] = []
                td["column_types"][str(x)] = "string"
            else:
                td["headers"].append(header.text)
                td["table"][header.text] = []
        first = True
        for row in table.find("tbody").find_all("tr"):
            if not (
                "style" in row.attrs
                and "display:none" in row.attrs["style"].replace(" ", "")
            ):
                for x, cell in enumerate(row.find_all("td")):
                    td["table"][td["headers"][x]].append(cell.text)
                    if first and "class" in cell.attrs:
                        td["column_types"][td["headers"][x]] = cell.attrs["class"][0]
                    elif first:
                        td["column_types"][td["headers"][x]] = "string"
                first = False
        if table.find("tfoot") and table.find("tfoot").find("tr"):
            for footer in table.find("tfoot").find("tr").find_all("th"):
                td["footers"].append(footer.text)
        return td

    def normalize_html(self):
        table = self.soup.find("table")
        header_trs = table.find_all("tr")
        while table.find(colspan=True) or table.find(rowspan=True):
            self.clean_spans()
        if not table.find("thead"):
            header_trs[0].wrap(self.soup.new_tag("thead"))
        for th in table.find("thead").find_all("td"):
            th.name = "th"
        if not table.find("tbody"):
            tbody_trs = table.find_all("tr", recursive=False)
            tbody_trs_list = []
            for tbody_tr in tbody_trs:
                tbody_trs_list.append(tbody_tr.extract())
            table.append(self.soup.new_tag("tbody"))
            for tbody_tr in tbody_trs_list:
                table.find("tbody").append(tbody_tr)
        for th in table.find("tbody").find_all("th"):
            th.name = "td"
        if table.find("tfoot"):
            for th in table.find("tfoot").find_all("td"):
                th.name = "th"

    def clean_spans(self):
        header_trs = self.soup.find("table").find_all("tr")
        for x, tr in enumerate(header_trs):
            cells = tr.select("th, td")
            for y, ele in enumerate(cells):
                if ele.has_attr("colspan"):
                    colspan = int(ele.attrs.pop("colspan"))
                    if colspan > 1:
                        new_tag = self.soup.new_tag(ele.name)
                        new_tag.string = ele.text
                        new_tag.attrs = ele.attrs
                        for _ in range(1, colspan):
                            tr.insert(y, copy.copy(new_tag))
                if ele.has_attr("rowspan"):
                    rowspan = int(ele.attrs.pop("rowspan"))
                    if rowspan > 1:
                        new_tag = self.soup.new_tag(ele.name)
                        new_tag.string = ele.text
                        new_tag.attrs = ele.attrs
                        for z in range(1, rowspan):
                            self.soup.find_all("tr")[x + z : x + z + 1][0].insert(
                                y, copy.copy(new_tag)
                            )


def open_xls_as_xlsx(filename):
    try:
        import xlrd
        from openpyxl.workbook import Workbook
        from openpyxl.reader.excel import load_workbook

        book = xlrd.open_workbook(filename)
        index = 0
        nrows, ncols = 0, 0
        sheet = None
        while nrows * ncols == 0:
            sheet = book.sheet_by_index(index)
            nrows = sheet.nrows
            ncols = sheet.ncols
            index += 1
        book1 = Workbook()
        sheet1 = book1.get_active_sheet()
        for row in range(0, nrows):
            for col in range(0, ncols):
                sheet1.cell(row=row + 1, column=col + 1).value = sheet.cell_value(
                    row, col
                )
        filename = filename.replace(".xls", ".xlsx")
        book1.save(filename)
        return Path(filename)
    except ImportError:
        print("xlrd and openpyxl are required in order to convert an xls file to xlsx")
        raise

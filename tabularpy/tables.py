import inspect
import locale
import re
from abc import ABCMeta
from collections import Counter
from copy import deepcopy
from csv import DictReader
from datetime import timedelta
from decimal import Decimal

from . import Settings
from .cell import Cell
from .col import Col
from .row import Row
from .util import (
    BeautifulSoupParser,
    clean_value,
    find_duplicates,
    get_sql_query_types,
    getindexes,
    getitem,
    merge_sort,
    transpose,
)

locale.setlocale(locale.LC_ALL, "")


class BaseTable(metaclass=ABCMeta):
    __slots__ = (
        "headers",
        "footers",
        "num_rows",
        "num_cols",
        "sql_table",
        "name",
        "_i",
        "_table_data",
        "column_types",
        "_settings",
    )

    def __init__(
        self, headers=None, footers=None, column_types=None, name=None, settings=None,
    ):
        self.headers = headers or []
        self.footers = footers or []
        self.column_types = column_types or {}
        self.num_rows = 0
        self.num_cols = 0
        self.sql_table = None
        self.name = name
        self._table_data = {
            "table": {},
            "headers": [],
            "footers": [],
            "column_types": {},
        }
        self._settings = settings or Settings()
        self._i = 0

    @classmethod
    def _setup(cls, obj):
        pass

    def _initialize(self):
        cf = self._settings.cell_formatter
        hf = self._settings.header_formatter
        table = self._table_data["table"]
        if not self.headers:
            if self._table_data["headers"]:
                self.headers = self._table_data["headers"]
            else:
                for x in range(len(table)):
                    self.headers.append(str(x))
        if table:
            self.num_rows = len(table[self.headers[0]])
        self.num_cols = len(self.headers)
        if not self.column_types:
            if self._table_data["column_types"]:
                self.column_types = self._table_data["column_types"]
            else:
                self.column_types = {header: None for header in self.headers}
        if not self.footers:
            if self._table_data["footers"]:
                self.footers = self._table_data["footers"]
        if hf:
            self.headers = [hf(header) for header in self.headers]
            self.column_types = {
                hf(key): self.column_types[key] for key in self.column_types
            }
            if table:
                table = {hf(key): table[key] for key in table}
        self._standardize_types()
        if table:
            self._table_data = table
        else:
            self._table_data = {header: [] for header in self.headers}
        if cf:
            self.format_cells(cf)
        if not all(list(self.column_types.values())):
            self.guess_types_from_data()

    @property
    def totals(self):
        totals = []
        for row in self:
            total = 0
            for cell in row:
                if cell.is_numeric():
                    total += cell.value
            totals.append(total)
        return totals

    def keys(self):
        return self.headers

    def reflect(self):
        for name, type_ in self.sql_table.get_column_types().items():
            self.column_types[name] = str(type_).lower()

    def add_column(self, column, name, col_type=None, prepend=False):
        if self._has_column(name):
            raise ValueError(f"{name} already exists in table")
        elif len(column) == self.num_rows:
            self._table_data[name] = column
            if prepend:
                self.headers.insert(0, name)
            else:
                self.headers.append(name)
            if col_type:
                self.column_types[name] = col_type
            else:
                self.column_types[name] = None
                if self.num_rows:
                    self.guess_types_from_data()
            self.num_cols += 1
        else:
            raise ValueError("shape of column does not match table")

    def add_row(self, row, prepend=False):
        if isinstance(row, dict):
            row_headers = row.keys()
            for row_header in row_headers:
                if row_header not in self.headers:
                    raise ValueError(f"{row} contains columns not in table")
            for header in self.headers:
                if not prepend:
                    try:
                        self._table_data[header].append(row[header])
                    except KeyError:
                        self._table_data[header].append(None)
                else:
                    try:
                        self._table_data[header].insert(0, row[header])
                    except KeyError:
                        self._table_data[header].insert(0, None)
            self.num_rows += 1
        elif isinstance(row, Row):
            if row.headers == self.headers:
                for cell in row:
                    self._table_data[cell.header].append(cell.raw_value)
                self.num_rows += 1
            else:
                raise AttributeError("Columns do not match")
        else:
            raise ValueError(f"{row} is not of type dict")

    def set_type(self, column, type_):
        self.column_types[column] = type_
        self._standardize_types()

    def format_cells(self, cell_formatter, replace=True):
        arg_len = len(inspect.signature(cell_formatter).parameters)
        if arg_len == 1:
            for key in self._table_data:
                for x, c in enumerate(self._table_data[key]):
                    self._table_data[key][x] = cell_formatter(c)
        elif arg_len == 2:
            for key in self._table_data:
                if self.column_types:
                    self._table_data[key] = cell_formatter(
                        self._table_data[key], self.column_types.get(key, None)
                    )
                else:
                    self._table_data[key] = cell_formatter(self._table_data[key], None)
        elif arg_len >= 3:
            for key in self._table_data:
                if self.column_types:
                    self._table_data[key] = cell_formatter(
                        self._table_data[key], self.column_types.get(key, None), key
                    )
                else:
                    self._table_data[key] = cell_formatter(
                        self._table_data[key], None, key
                    )
        if replace:
            self._settings.cell_formatter = cell_formatter

    def format_column(self, column, formatter):
        arg_len = len(inspect.signature(formatter).parameters)
        if arg_len == 1:
            for x, c in enumerate(self._table_data[column]):
                self._table_data[column][x] = formatter(c)
        elif arg_len == 2:
            if self._has_column(column):
                if self.column_types:
                    self._table_data[column] = formatter(
                        self._table_data[column], self.column_types.get(column, None)
                    )
                else:
                    self._table_data[column] = formatter(self._table_data[column], None)
        elif arg_len >= 3:
            if self._has_column(column):
                if self.column_types:
                    self._table_data[column] = formatter(
                        self._table_data[column],
                        self.column_types.get(column, None),
                        column,
                    )
                else:
                    self._table_data[column] = formatter(
                        self._table_data[column], None, column
                    )
        return self

    def format_headers(self, formatter):
        for x in range(len(self.headers)):
            self._table_data[formatter(self.headers[x])] = self._table_data.pop(
                self.headers[x]
            )
            self.column_types[formatter(self.headers[x])] = self.column_types.pop(
                self.headers[x]
            )
            self.headers[x] = formatter(self.headers[x])

    def _standardize_types(self):
        for key, datatype in self.column_types.items():
            if isinstance(datatype, dict):
                if "class" in datatype:
                    datatype = datatype["class"]
            datatype = str(datatype).lower()
            if "interv" in datatype:
                self.column_types[key] = "interval"
            elif "dec" in datatype or "flo" in datatype or "numeric" in datatype:
                self.column_types[key] = "numeric"
            elif "int" in datatype or "num" in datatype:
                self.column_types[key] = "integer"
            elif (
                "varchar" in datatype
                or "character var" in datatype
                or "str" in datatype
                or "text" in datatype
            ):
                self.column_types[key] = "varchar"
            elif "datetimesec" in datatype or "second" in datatype:
                self.column_types[key] = "seconds"
            elif "date" in datatype:
                self.column_types[key] = "date"
            elif "times" in datatype:
                if key == "month":
                    self.column_types[key] = "date"
                else:
                    self.column_types[key] = "timestamp"
            elif "time" in datatype:
                self.column_types[key] = "time"
            elif "perc" in datatype:
                self.column_types[key] = "percent"
            elif "month" in datatype:
                self.column_types[key] = "date"
            elif "mon" in datatype or "curenc" in datatype:
                self.column_types[key] = "money"
            elif "boo" in datatype:
                self.column_types[key] = "bool"

    def guess_types_from_data(self, guess_function=None):
        if guess_function:
            self.column_types = guess_function(self)
        else:
            if not self._settings.do_not_guess_types:
                for cell in self[0]:
                    if self.column_types[cell.header] is None:
                        val = str(cell.value)
                        if ("/" in val or "-" in val) and ":" in val:
                            self.column_types[cell.header] = "timestamp"
                        elif (
                            ("/" in val or "-" in val)
                            and not re.search("[a-zA-Z]", val)
                            and (len(val) == 10 or len(val) == 8)
                            and val[0] != "-"
                        ):
                            self.column_types[cell.header] = "date"
                        elif ":" in val:
                            self.column_types[cell.header] = "time"
                        elif "$" in val:
                            self.column_types[cell.header] = "money"
                        elif "%" in val:
                            self.column_types[cell.header] = "percent"
                        elif (
                            "." in val
                            and val.replace(".", "").replace("-", "").isdecimal()
                            and val.count(".") == 1
                        ):
                            self.column_types[cell.header] = "numeric"
                        elif val.replace("-", "").isnumeric() and (
                            val.count("-") == 0
                            or (val.count("-") == 1 and val[1] == "-")
                        ):
                            type_string = "integer"
                            col = self[cell.header]
                            for col_cell in col:
                                if col_cell.value is not None:
                                    col_cell_val = str(col_cell.value)
                                    if (
                                        "." in col_cell_val
                                        and col_cell_val.replace(".", "")
                                        .replace("-", "")
                                        .isdecimal()
                                        and col_cell_val.count(".") == 1
                                    ):
                                        type_string = "numeric"
                                        break
                                    elif (
                                        int(col_cell_val.replace(",", "")) < -2147483648
                                        or int(col_cell_val.replace(",", ""))
                                        > 2147483647
                                    ):
                                        type_string = "bigint"
                                        break
                            self.column_types[cell.header] = type_string
                        else:
                            self.column_types[cell.header] = "varchar"
                            pass

    def rename_column(self, old_column, new_column):
        if self._has_column(old_column):
            self._table_data[new_column] = self._table_data.pop(old_column)
            self.headers[self.headers.index(old_column)] = new_column
            if old_column in self.column_types:
                self.column_types[new_column] = self.column_types.pop(old_column)
        return self

    def delete(self, *args):
        ints = []
        for item in args:
            if isinstance(item, str):
                if item in self.headers:
                    del self.headers[self.headers.index(item)]
                    del self._table_data[item]
                if item in self.column_types:
                    del self.column_types[item]
            elif isinstance(item, int):
                ints.append(item)
        ints.sort()
        for item in reversed(ints):
            if item < self.num_rows:
                for key in self._table_data:
                    del self._table_data[key][item]
                self.num_rows -= 1
        return self

    def delete_by_column(self, column, *args):
        for item in args:
            idx = self[column].index(item)
            for key in self._table_data:
                del self._table_data[key][idx]
            self.num_rows -= 1
        return self

    def cell(self, x, y):
        if isinstance(y, str) and isinstance(x, int):
            if self._has_column(y):
                if self._has_row(x):
                    return Cell(
                        self._table_data[y][x], y, x, self.headers.index(y), self
                    )
                else:
                    raise AttributeError(f"{self} does not have row {x}")
            else:
                raise AttributeError(f"{self} does not have column {y}")
        elif isinstance(y, int) and isinstance(x, int):
            header = self.headers[y]
            return Cell(self._table_data[header][x], header, x, y, self._get_row(x))

    def change_cell(self, x, y, value):
        if isinstance(y, str) and isinstance(x, int):
            if self._has_column(y):
                if self._has_row(x):
                    self._table_data[y][x] = value
                else:
                    raise AttributeError(f"{self} does not have row {x}")
            else:
                raise AttributeError(f"{self} does not have column {y}")
        elif isinstance(y, int) and isinstance(x, int):
            header = self.headers[y]
            if self._has_row(x):
                self._table_data[header][x] = value
            else:
                raise AttributeError(f"{self} does not have row {x}")

    def count_val(self, value, column=None):
        val_count = 0
        if self:
            if column:
                for cell in self._table_data[column]:
                    if value == cell:
                        val_count += 1
            else:
                for column in self.headers:
                    for cell in self._table_data[column]:
                        if value == cell:
                            val_count += 1
        return val_count

    def copy(self, without_data=False):
        if not without_data:
            lists = self.to_list_of_lists()
            if lists:
                return ListOfListsTable(
                    lists,
                    self.headers.copy(),
                    self.footers.copy(),
                    self.column_types.copy(),
                    self.name,
                    deepcopy(self._settings),
                )
        return Table(
            self.headers.copy(),
            self.footers.copy(),
            self.column_types.copy(),
            self.name,
            deepcopy(self._settings),
        )

    def remove_duplicates(self, *columns):
        dup_rows = self.find_duplicates(*columns)
        for row in reversed(dup_rows):
            del self[row]

    def find_duplicates(self, *columns):
        if self:
            if not columns:
                columns = self.headers
            column_rows = []
            for column in columns:
                if self._has_column(column):
                    column_rows.append(self._table_data[column])
            if not column_rows:
                return
            column_rows = transpose(column_rows)
            column_rows = [" ".join([str(item) for item in row]) for row in column_rows]
            results = find_duplicates(column_rows)
            dup_rows = set()
            for key in results:
                results[key].pop(0)
                for row in results[key]:
                    dup_rows.add(row)
            return sorted(dup_rows)

    def pop_row(self, where):
        """The 'where' argument can be a number which represents a row or it
        can be a dictionary with keys representing columns and the values
        representing the table values for the matching keys. The first row
        with matching values in the specified columns will be removed from
        the table and returned from the function"""
        if isinstance(where, int):
            if self._has_row(where):
                row = self[where]
                r = {
                    cell.header: self._table_data[cell.header].pop(where)
                    for cell in row
                }
                self.num_rows -= 1
                return r
        else:
            indecies = []
            for key, value in where.items():
                indecies.append({i for i, x in enumerate(self[key]) if x == value})
            row = set.intersection(*indecies)
            r = {
                cell.header: self._table_data[cell.header].pop()
                for cell in self[list(row)[0]]
            }
            self.num_rows -= 1
            return r

    def pop_column(self, col_name):
        if isinstance(col_name, str):
            if self._has_column(col_name):
                c = self._table_data.pop(col_name)
                self.headers.remove(col_name)
                self.num_cols -= 1
                return c
        elif isinstance(col_name, int):
            if col_name <= self.num_cols:
                pass

    def filter(self, column, condition, what):
        if condition == "==":
            condition = lambda x, y: x == y
        elif condition == "<=":
            condition = lambda x, y: x <= y
        elif condition == ">=":
            condition = lambda x, y: x >= y
        elif condition == ">":
            condition = lambda x, y: x > y
        elif condition == "<":
            condition = lambda x, y: x < y
        elif condition == "!=":
            condition = lambda x, y: x != y
        else:
            raise SyntaxError
        filtered_table = Table(
            self.headers, self.footers, self.column_types, self.name, self._settings
        )
        for row in self:
            if condition(row[column], what):
                filtered_table.add_row(row)
        return filtered_table

    def columns(self):
        return [self._get_column(header) for header in self.headers]

    def field_definitions(self):
        output = []
        for header in self.headers:
            output.append({"name": header, "type": self.column_types[header]})
        return output

    def replace(self, ident, old, new):
        if isinstance(ident, str):
            if self._has_column(ident):
                self._get_column(ident).replace(old, new)
        elif isinstance(ident, int):
            if self._has_row(ident):
                self._get_row(ident).replace(old, new)

    def replace_in_column(self, col, key):
        if self._has_column(col):
            for row in self:
                self._table_data[col][row.row_num] = key(row[col]._value)

    def pprint(self, num_rows=None):
        col_lengths = [0 for _ in range(self.num_cols)]
        if num_rows:
            row_num_length = len(str(num_rows + 1))
        else:
            row_num_length = len(str(self.num_rows + 1))
        if row_num_length < 4:
            row_num_length = 4
        for x, header in enumerate(self.headers):
            if len(str(header)) > col_lengths[x]:
                col_lengths[x] = len(str(header))
        for i, row in enumerate(self):
            for x, cell in enumerate(row):
                if len(str(cell)) > col_lengths[x]:
                    col_lengths[x] = len(str(cell))
            if num_rows and i >= num_rows:
                break
        table_string = f"{'row':>{row_num_length}}"
        for x, header in enumerate(self.headers):
            table_string = f"{table_string}\t{header:>{col_lengths[x]}}"
        table_string = f"{table_string}\r\n"
        for i, row in enumerate(self, 1):
            table_string = f"{table_string}{i:>{row_num_length}}"
            for x, cell in enumerate(row):
                table_string = f"{table_string}\t{str(cell):>{col_lengths[x]}}"
            table_string = f"{table_string}\r\n"
            if num_rows and i >= num_rows:
                break
        return table_string

    def to_dicts(self, *columns):
        if columns:
            return ({col: row[col] for col in columns} for row in self)
        else:
            return (row.to_dict() for row in self)

    def to_list_of_dicts(self, *columns):
        if columns:
            return [{col: row[col] for col in columns} for row in self]
        else:
            return [row.to_dict() for row in self]

    def to_list_of_tuples(self, *columns):
        if columns:
            return [tuple(row[col] for col in columns) for row in self]
        else:
            return [row.to_tuple() for row in self]

    def to_list_of_lists(self, *columns, include_header=False):
        output = []
        if include_header:
            if columns:
                output.append(columns)
            else:
                output.append(self.headers)
        for row in self:
            if columns:
                output.append([row[col] for col in columns])
            else:
                output.append(row.to_list())
        return output

    def to_dict(self, row_num):
        return self._get_row(row_num).to_dict()

    def pivot(
        self,
        rows=None,
        columns=None,
        values=None,
        aggr_funcs=None,
        row_sort=None,
        col_sort=None,
    ):
        if self:
            if not isinstance(values, list):
                raise TypeError(
                    'The "values" parameter must be a list of value column headers'
                )
            r = []
            distinct_row = []
            row_indexes = {}
            c = []
            distinct_col = []
            col_indexes = {}
            column_types = {}

            if rows:
                r = self._table_data[rows]
                distinct_row = list(set(r))
            if columns:
                c = self._table_data[columns]
                distinct_col = list(set(c))
            vs = []
            for value in values:
                vs.append(
                    [
                        clean_value(val, self.column_types[value], self._settings)
                        for val in self._table_data[value]
                    ]
                )
            if distinct_row:
                if row_sort:
                    distinct_row.sort(key=row_sort)
                else:
                    distinct_row.sort()
                row_indexes = {val: getindexes(r, val) for val in distinct_row}
            if distinct_col:
                if col_sort:
                    distinct_col.sort(key=col_sort)
                else:
                    distinct_col.sort()
                col_indexes = {val: getindexes(c, val) for val in distinct_col}

            list_of_lists = []

            if row_indexes and col_indexes:
                for y, row_val in enumerate(distinct_row):
                    list_of_lists.append([row_val])
                    for x, col_val in enumerate(distinct_col):
                        intersection = set(row_indexes[row_val]).intersection(
                            col_indexes[col_val]
                        )
                        if intersection:
                            list_of_lists[y].append(0)
                            for n, i in enumerate(intersection):
                                # must fix to go through all the values and aggregate
                                # functions insead of just the first ones
                                list_of_lists[y][x + 1] = aggr_funcs(
                                    list_of_lists[y][x + 1], vs[0][i], n + 1
                                )
                        else:
                            list_of_lists[y].append(None)
                column_types = {col: self.column_types[values] for col in distinct_col}
                column_types[rows] = self.column_types[rows]
                distinct_col.insert(0, rows)

            elif row_indexes:
                for y, (row_val, n) in enumerate(row_indexes.items()):
                    list_of_lists.append([row_val])
                    for x, v in enumerate(vs):
                        list_of_lists[y].append(0)
                        for i, row in enumerate(n):
                            list_of_lists[y][x + 1] = aggr_funcs[x](
                                list_of_lists[y][x + 1], v[row], i + 1
                            )
                distinct_col = [rows, *values]
                column_types = {value: self.column_types[value] for value in values}
                column_types[rows] = self.column_types[rows]

            # Haven't tried this one yet, probably doesn't work as expected!
            elif col_indexes:
                for y, (col_val, n) in enumerate(col_indexes.items()):
                    list_of_lists.append([col_val])
                    for x, v in enumerate(vs):
                        list_of_lists[y].append(0)
                        for i, row in enumerate(n):
                            list_of_lists[y][x + 1] = aggr_funcs[x](
                                list_of_lists[y][x + 1], v[row], i + 1
                            )
                distinct_col = [columns, values]
                column_types = {
                    columns: self.column_types[columns],
                    values: self.column_types[values],
                }
                # Should have to transpose the headers (distinct_col)
                # and column_types as well
                transpose(list_of_lists)

            return ListOfListsTable(
                list_of_lists,
                distinct_col,
                self.footers,
                column_types,
                self.name,
                self._settings,
            )

    def to_excel(self, formatter=None):
        try:
            # noinspection PyUnresolvedReferences
            import openpyxl

            # noinspection PyUnresolvedReferences
            from openpyxl.styles import Font, Alignment
        except ImportError:
            print(
                "openpyxl is required in order to create an Excel file. Alternativly",
                "one could create a csv file using .to_csv(), which is compatible with "
                "Excel.",
            )
            raise
        wb = openpyxl.Workbook()
        ws = wb.active
        for y, cell in enumerate(self.headers, start=1):
            ws.cell(column=y, row=1).value = cell
            ws.cell(column=y, row=1).font = Font(bold=True)
            ws.cell(column=y, row=1).alignment = Alignment(
                wrap_text=True, horizontal="center", vertical="center"
            )
        ws.row_dimensions[1].height = 30
        try:
            for y, row in enumerate(self, start=2):
                for x, cell in enumerate(row, start=1):
                    header = self.headers[x - 1]
                    if formatter:
                        form = formatter(self.column_types[header], header)
                        if form:
                            ws.cell(column=x, row=y).number_format = form
                    ws.cell(column=x, row=y).value = cell.value
                    ws.cell(column=x, row=y).alignment = Alignment(horizontal="right")
        except IndexError:
            pass
        lastrow = len(self) + 2
        if self.footers:
            for y, cell in enumerate(self.footers, start=1):
                ws.cell(column=y, row=lastrow).value = cell
                ws.cell(column=y, row=lastrow).number_format = ws.cell(
                    column=y, row=lastrow - 1
                ).number_format
                ws.cell(column=y, row=lastrow).font = Font(bold=True)
                ws.cell(column=y, row=lastrow).alignment = Alignment(horizontal="right")
        return wb

    def to_csv(self, header=True, footer=False, handle_none=False):
        csv_pieces = []
        if self.headers and header:
            for head in self.headers:
                head = str(head)
                if "," in head:
                    csv_pieces.append(f'"{head}",')
                else:
                    csv_pieces.append(f"{head},")
            csv_pieces = "".join(csv_pieces)
            csv_pieces = [f"{csv_pieces[:-1]}\n"]
        for row in self:
            for cell in row:
                if handle_none and cell.value is None:
                    cell = ""
                cell = str(cell)
                if "," in cell:
                    csv_pieces.append(f'"{cell}",')
                else:
                    csv_pieces.append(f"{cell},")
            csv_pieces = "".join(csv_pieces)
            csv_pieces = [f"{csv_pieces[:-1]}\n"]
        if self.footers and footer:
            for footer in self.footers:
                footer = str(footer)
                if "," in footer:
                    csv_pieces.append(f'"{footer}",')
                else:
                    csv_pieces.append(f"{footer},")
            csv_pieces = "".join(csv_pieces)
            csv_pieces = [f"{csv_pieces[:-1]}\n"]
        return "".join(csv_pieces)

    def to_json_string(self, json_type="array_of_objects"):
        json_pieces = []
        if json_type == "array_of_objects":
            if self.headers:
                json_pieces.append("[")
                for row in self:
                    json_pieces.append("{")
                    if self.column_types:
                        for header in self.headers:
                            cell = row[header]
                            if (
                                "int" in str(cell.column_type).lower()
                                or "float" in str(cell.column_type).lower()
                                or "decimal" in str(cell.column_type).lower()
                            ):
                                json_pieces.append(
                                    f'"{str(header)}":{str(cell).replace(",", "")}, '
                                )
                            else:
                                json_pieces.append(f'"{str(header)}":"{str(cell)}", ')
                    else:
                        for header in self.headers:
                            json_pieces.append(
                                f'"{str(header)}":"{str(row[header])}", '
                            )
                    json_pieces = "".join(json_pieces)
                    json_pieces = [f"{json_pieces[:-2]}}}, "]
                json_pieces = "".join(json_pieces)
                json_pieces = [f"{json_pieces[:-2]}]"]
        elif json_type == "array_of_arrays":
            json_pieces.append("[")
            if self.headers:
                json_pieces.append("[")
                for header in self.headers:
                    json_pieces.append(f'"{str(header)}", ')
                json_pieces = "".join(json_pieces)
                json_pieces = [f"{json_pieces[:-2]}], "]
            for row in self:
                json_pieces.append("[")
                if self.column_types:
                    for cell in row:
                        if (
                            "int" in str(cell.column_type).lower()
                            or "float" in str(cell.column_type).lower()
                            or "decimal" in str(cell.column_type).lower()
                        ):
                            json_pieces.append('{str(cell).replace(",", "")}, ')
                        else:
                            json_pieces.append(f'"{str(cell)}", ')
                else:
                    for cell in row:
                        json_pieces.append(f'"{str(cell)}", ')
                json_pieces = "".join(json_pieces)
                json_pieces = [f"{json_pieces[:-2]}}}, "]
            json_pieces = "".join(json_pieces)
            json_pieces = [f"{json_pieces[:-2]}]"]
        return "".join(json_pieces)

    def to_html_table(
        self,
        inner_table=False,
        footer=False,
        full_html=False,
        add_attr=None,
        table_attr=None,
        header_formatter=None,
        row_totals=False,
        col_totals=False,
        col_group=False,
    ):
        html_pieces = []
        if full_html:
            html_pieces.append("<html><body>")
        if not inner_table:
            if table_attr:
                html_pieces.append(f"<table {table_attr}>")
            else:
                html_pieces.append("<table>")
        if col_group:
            html_pieces.append("<colgroup>")
            for _ in self.headers:
                html_pieces.append("<col>")
            html_pieces.append("</colgroup>")
        if self.headers:
            html_pieces.append("<thead><tr>")
            html_pieces.append(f"{self.headers_to_html(header_formatter, add_attr)}")
            if row_totals:
                html_pieces.append("<th>Total</th>")
            html_pieces.append("</tr></thead>")
        if self.footers:
            html_pieces.append("<tfoot><tr>")
            html_pieces.append(f"{self.footers_to_html()}")
            html_pieces.append("</tr></tfoot>")
        elif footer:
            html_pieces.append("<tfoot></tfoot>")
        html_pieces.append("<tbody>")
        grand_total = 0
        for y, row in enumerate(self):
            html_pieces.append(f"<tr>{row.to_html(add_attr, row_totals)}</tr>")
            if row_totals:
                grand_total += row.sum()
        if col_totals:
            grand_total = 0
            html_pieces.append("<tr>")
            for col in self.columns():
                total = col.sum()
                html_pieces.append(f'<td class="coltotal">{total:,}</td>')
                grand_total += total
            if row_totals:
                html_pieces.append(f'<td class="grandtotal">{grand_total:,}</td>')
            html_pieces.append("</tr>")
        html_pieces.append("</tbody>")
        if not inner_table:
            html_pieces.append("</table>")
        if full_html:
            html_pieces.append("</body></html>")
        return "".join(html_pieces)

    def headers_to_html(self, header_formatter=None, add_attr=None):
        if header_formatter:
            if add_attr:
                html = "".join(
                    f"<th {add_attr(header, self.column_types[header], header)}>"
                    f"{header_formatter(header)}</th>"
                    for header in self.headers
                )
            else:
                html = "".join(
                    f"<th>{header_formatter(header)}</th>" for header in self.headers
                )
        else:
            if add_attr:
                html = "".join(
                    f"<th {add_attr(header, self.column_types[header], header)}>"
                    f"{header}</th>"
                    for header in self.headers
                )
            else:
                html = "".join(f"<th>{header}</th>" for header in self.headers)
        return html

    def footers_to_html(self):
        return "".join([f"<th>{footer}</th>" for footer in self.footers])

    def sort(self, header, key=None, reverse=False):
        if self._has_column(header):
            idx = self.headers.index(header)
            ziped_values = list(
                zip(*[self._table_data[column] for column in self.headers])
            )
            if isinstance(key, dict):
                key_dict = key
                key = lambda x: key_dict[x]
            ziped_values = merge_sort(ziped_values, idx, key, reverse)
            ziped_values = transpose(ziped_values)
            for header, values in zip(self.headers, ziped_values):
                self._table_data[header] = list(values)

    def _has_column(self, col):
        if col in self.headers:
            return True
        return False

    def _get_column(self, header):
        if self._has_column(header):
            return Col(
                self._table_data[header],
                self.column_types[header],
                header,
                self.headers.index(header),
                self,
                self._settings,
            )
        else:
            raise AttributeError(f"{self} does not have column {header}")

    def _has_row(self, row_num):
        if row_num < self.num_rows:
            return True
        return False

    def _get_row(self, row_num):
        if self._has_row(row_num):
            row = [self._table_data[col][row_num] for col in self.headers]
            return Row(
                row,
                list(self.headers),
                self.column_types.copy(),
                row_num,
                self,
                self._settings,
            )

    def __getattr__(self, item):
        if self._has_column(item):
            return self._get_column(item)

    def __getitem__(self, args):
        if isinstance(args, tuple):
            select = []
            if isinstance(args[0], str):
                for item in args:
                    select.append(self._get_column(item).to_list())
                return ListOfListsTable(
                    transpose(select),
                    args,
                    self.footers,
                    {header: self.column_types[header] for header in args},
                    self.name,
                    self._settings,
                )
            elif isinstance(args[0], int):
                for item in args:
                    select.append(self._get_row(item).to_list())
                return ListOfListsTable(
                    select,
                    self.headers,
                    self.footers,
                    self.column_types,
                    self.name,
                    self._settings,
                )
        else:
            if isinstance(args, slice):
                select_rows = []
                for x in range(args.start, args.stop, args.step):
                    select_rows.append(self._get_row(x).to_list())
                return ListOfListsTable(
                    select_rows,
                    self.headers,
                    self.footers,
                    self.column_types,
                    self.name,
                    self._settings,
                )
            elif isinstance(args, str):
                return self._get_column(args)
            elif isinstance(args, int):
                return self._get_row(args)

    def __setitem__(self, key, value):
        if isinstance(key, str):
            if isinstance(value, list):
                if len(value) == self.num_rows:
                    self._table_data[key] = value
                elif len(value) == 0:
                    self._table_data[key] = [None for _ in range(self.num_rows)]
                else:
                    raise ValueError(
                        f"number of items in provided list, {len(value)}, "
                        f"does not match objects rows, {self.num_rows}"
                    )
                self.num_cols += 1
                if key not in self.headers:
                    self.headers.append(key)
            else:
                raise ValueError(f"{type(value)} is not of type list")
        else:
            raise ValueError(f"{type(key)} is not of type string")

    def __delitem__(self, item):
        if isinstance(item, str):
            if self._has_column(item):
                formatted_item = item
                del self._table_data[formatted_item]
                self.headers.remove(formatted_item)
                if item in self.column_types:
                    del self.column_types[item]
                if self.num_cols > 0:
                    self.num_cols -= 1
        elif isinstance(item, int):
            if item < self.num_rows:
                for col in self._table_data:
                    del self._table_data[col][item]
                self.num_rows -= 1
            else:
                raise IndexError(f"{item} is out of range")

    def __contains__(self, item):
        if isinstance(item, str):
            if (
                self._settings.header_formatter
                and self._settings.header_formatter(item) in self.headers
            ):
                return True
            elif item in self.headers:
                return True
            else:
                return False
        elif isinstance(item, int):
            if self._has_row(item):
                return True
            else:
                return False

    def __iter__(self):
        return self

    def __next__(self):
        if self._i < self.num_rows:
            row = self._get_row(self._i)
            self._i += 1
            return row
        else:
            self._i = 0
            raise StopIteration

    def __len__(self):
        return self.num_rows

    def __repr__(self):
        if self.name:
            return f"<Table(name={self.name})>"
        else:
            return "<Table()>"

    def __str__(self):
        return "\n".join(str(row) for row in self)

    def __bool__(self):
        if self.num_cols and self.num_rows:
            return True
        else:
            return False

    def __add__(self, other):
        if isinstance(other, BaseTable):
            if self.headers == other.headers:
                if self and other:
                    list1 = self.to_list_of_lists()
                    list2 = other.to_list_of_lists()
                    list1.extend(list2)
                    return ListOfListsTable(
                        list1,
                        self.headers,
                        self.footers,
                        self.column_types,
                        self.name,
                        self._settings,
                    )
                elif not self and other:
                    return other.copy()
                elif not self and not other:
                    return Table(
                        self.headers,
                        self.footers,
                        self.column_types,
                        self.name,
                        self._settings,
                    )
                else:
                    return self.copy()
            else:
                raise ValueError("Columns do not match")
        elif other is None:
            return self.copy()
        else:
            raise TypeError(
                f"unsupported operand type(s) for +: 'Table' and {type(other)}"
            )

    def __sub__(self, other):
        if isinstance(other, BaseTable):
            if self.headers == other.headers:
                if self and other:
                    self_copy = self.copy()
                    self_set = set(self.to_list_of_tuples())
                    other_set = set(other.to_list_of_tuples())
                    intersections = self_set.intersection(other_set)
                    for row in self_copy:
                        if row.to_tuple() in intersections:
                            del self_copy[row.row_num]
                    if self_copy:
                        return self_copy
                else:
                    return self.copy()
            else:
                raise ValueError(
                    "Columns of each object do not match in either length or order"
                )
        else:
            raise TypeError(
                f"unsupported operand type(s) for -: 'Table' and {type(other)}"
            )

    def __eq__(self, other):
        if not isinstance(other, BaseTable):
            return False
        if self._table_data != other._table_data:
            return False
        if self.column_types != other.column_types:
            return False
        if self.footers != other.footers:
            return False
        if self.header != other.header:
            return False
        if self.num_cols != other.num_cols:
            return False
        if self.num_rows != other.num_rows:
            return False
        return True

    def __ne__(self, other):
        return not self.__eq__(other)


class CsvTable(BaseTable):
    def __init__(
        self,
        file_path,
        delimiter=",",
        headers=None,
        footers=None,
        column_types=None,
        name=None,
        settings=None,
    ):
        super().__init__(headers, footers, column_types, name, settings)
        self.delimiter = delimiter
        try:
            with open(file_path) as open_file:
                self._setup(open_file)
        except (TypeError, OSError):
            self._setup(file_path.split("\n"))

    def _setup(self, obj):
        if self.headers:
            temp = DictReader(obj, delimiter=self.delimiter, fieldnames=self.headers)
            self._read_file(temp)
        else:
            temp = DictReader(obj, delimiter=self.delimiter)
            self._read_file(temp)

    def _read_file(self, obj):
        if obj:
            if self.check_duplicates(obj.fieldnames):
                raise ValueError("duplicate headers")
            self._table_data["headers"] = list(obj.fieldnames)
            obj = list(obj)
            for x, row in enumerate(obj):
                obj[x] = [row[header] for header in self._table_data["headers"]]
            if self.is_empty(obj):
                return
            obj = transpose(obj)
            for header, column in zip(self._table_data["headers"], obj):
                self._table_data["table"][header] = column
        self._initialize()

    @staticmethod
    def check_duplicates(headers):
        return bool([k for k, v in Counter(headers).items() if v > 1])

    @staticmethod
    def is_empty(obj):
        if not obj:
            return True
        temp = [a for a in obj[0] if a is ""]
        if len(temp) == len(obj[0]):
            return True
        return False

    def __eq__(self, other):
        return super().__eq__(other)


class ExcelTable(BaseTable):
    def __init__(
        self,
        file_path,
        worksheet=0,
        headers=None,
        footers=None,
        column_types=None,
        name=None,
        settings=None,
    ):
        super().__init__(headers, footers, column_types, name, settings)
        self._setup(file_path)
        self.worksheet = worksheet

    def _setup(self, obj):
        try:
            # noinspection PyUnresolvedReferences
            from openpyxl import load_workbook

            # noinspection PyUnresolvedReferences
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("openpyxl is required in order to read an Excel file")
            raise
        wb = load_workbook(filename=str(obj))
        if isinstance(self.worksheet, int):
            ws = wb[wb.get_sheet_names()[self.worksheet]]
        elif isinstance(self.worksheet, str):
            ws = wb[self.worksheet]
        else:
            ws = wb[wb.get_sheet_names()[0]]
        if not self.headers:
            for cell in list(ws[f"A1:{get_column_letter(ws.max_column)}1"])[0]:
                self._table_data["headers"].append(cell.value)
        if not self.column_types:
            self._guess_excel_types(
                list(ws[f"A2:{get_column_letter(ws.max_column)}2"])[0]
            )
        for col, header in zip(ws.columns, self._table_data["headers"]):
            first = True
            self._table_data["table"][header] = []
            for cell in col:
                if not first:
                    if self._table_data["column_types"][header] == "interval":
                        self._table_data["table"][header].append(
                            timedelta(days=cell.value * 24)
                        )
                    elif self._table_data["column_types"][header] == "decimal":
                        self._table_data["table"][header].append(
                            Decimal(str(cell.value))
                        )
                    else:
                        self._table_data["table"][header].append(cell.value)
                else:
                    first = False
        self._initialize()

    def _guess_excel_types(self, row):
        for cell, header in zip(row, self._table_data["headers"]):
            if cell.value is not None:
                number_format = cell.number_format
                data_type = cell.data_type
                val = str(cell.value)
                if (
                    number_format
                    == '_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)'
                    or number_format == '"$"#,##0.00'
                ):
                    self._table_data["column_types"][header] = "money"
                elif number_format == "0%":
                    self._table_data["column_types"][header] = "percent"
                elif number_format == "0.00":
                    self._table_data["column_types"][header] = "decimal"
                elif number_format == "[h]:mm:ss;@":
                    self._table_data["column_types"][header] = "interval"
                elif (
                    number_format == "[$-409]h:mm:ss\ AM/PM;@"
                    or number_format == "[$-409]h:mm\ AM/PM;@"
                    or number_format == "h:mm:ss;@"
                    or number_format == "h:mm;@"
                ):
                    self._table_data["column_types"][header] = "time"
                elif (
                    number_format == "mm-dd-yy"
                    or number_format == "mm/dd/yy"
                    or number_format == "mm-dd-yyyy"
                    or number_format == "mm/dd/yyyy"
                    or number_format == "mm/yyyy"
                    or number_format == "mm-yyyy"
                    or number_format == "mm/yy"
                    or number_format == "mm-yy"
                ):
                    self._table_data["column_types"][header] = "date"
                elif ("/" in val or "-" in val) and ":" in val:
                    self._table_data["column_types"][header] = "timestamp"
                elif ("/" in val or "-" in val) and not re.search("[a-zA-Z]", val):
                    self._table_data["column_types"][header] = "date"
                elif number_format == "General" and data_type == "n":
                    self._table_data["column_types"][header] = "integer"
                else:
                    self._table_data["column_types"][header] = "string"
            else:
                self._table_data["column_types"][header] = "string"


class SqlAlcTable(BaseTable):
    __slots__ = "query"

    def __init__(
        self,
        result,
        query=None,
        footers=None,
        column_types=None,
        name=None,
        settings=None,
    ):
        super().__init__(None, footers, column_types, name, settings)
        self.query = query
        self._setup(result)

    def _setup(self, obj):
        if not isinstance(obj, list):
            obj = [obj]
        if self.query and not self.column_types:
            self._table_data["column_types"] = get_sql_query_types(self.query)
            self._table_data["headers"] = list(self._table_data["column_types"].keys())
        elif self.column_types:
            self._table_data["headers"] = list(self.column_types.keys())
        else:
            if obj:
                self._table_data["headers"] = [str(key) for key in obj[0].keys()]
        if obj:
            obj = transpose(obj)
            for header, column in zip(self._table_data["headers"], obj):
                self._table_data["table"][header] = column
        self._initialize()


class HtmlTable(BaseTable):
    __slots__ = "_parser"

    def __init__(
        self,
        html,
        parser=None,
        footers=None,
        column_types=None,
        name=None,
        settings=None,
    ):
        super().__init__(None, footers, column_types, name, settings)
        self._parser = parser or BeautifulSoupParser().parse
        self._setup(html)

    def _setup(self, html):
        self._table_data = self._parser(html)
        self._initialize()


class ListOfListsTable(BaseTable):
    def __init__(
        self,
        lists,
        headers=None,
        footers=None,
        column_types=None,
        name=None,
        settings=None,
    ):
        super().__init__(headers, footers, column_types, name, settings)
        self._setup(lists)

    def _setup(self, obj):
        if obj:
            obj = transpose(obj)
            for x, column in enumerate(obj):
                header = str(getitem(self.headers, x, x))
                self._table_data["table"][header] = column
            self._initialize()


class ListOfDictsTable(BaseTable):
    def __init__(
        self, dicts, footers=None, column_types=None, name=None, settings=None
    ):
        super().__init__(None, footers, column_types, name, settings)
        self._setup(dicts)

    def _setup(self, obj):
        if obj:
            if self._settings.sort_headers:
                self._table_data["headers"] = list(sorted(obj[0].keys()))
            else:
                self._table_data["headers"] = list(obj[0].keys())
            for x, row in enumerate(obj):
                obj[x] = [row[header] for header in self._table_data["headers"]]
            obj = transpose(obj)
            for header, column in zip(self._table_data["headers"], obj):
                self._table_data["table"][header] = column
            self._initialize()


LXML_TYPE_MAP = {
    "IntElement": "integer",
    "StringElement": "varchar",
    "BoolElement": "bool",
}


class LxmlTable(BaseTable):
    def __init__(
        self, xml_obj, pyval=False, name=None, settings=None,
    ):
        super().__init__(None, None, None, name, settings)
        self._setup(xml_obj)
        self._pyval = pyval

    def _setup(self, xml_obj):
        headers = []
        column_types = {}
        if len(xml_obj):
            for child in xml_obj[0].getchildren():
                if child.tag not in headers:
                    headers.append(child.tag)
                    child_type = (
                        str(type(child))
                        .replace("<class 'lxml.objectify.", "")
                        .replace("'>", "")
                    )
                    column_types[child.tag] = LXML_TYPE_MAP[child_type]
        self._table_data["headers"] = headers
        self._table_data["column_types"] = column_types
        for row in xml_obj:
            for header in headers:
                if header not in self._table_data["table"]:
                    self._table_data["table"][header] = []
                try:
                    if self._pyval:
                        value = row[header].pyval
                    else:
                        value = row[header].text
                except AttributeError:
                    self._table_data["table"][header].append("")
                else:
                    self._table_data["table"][header].append(value)
        self._initialize()


class Table(BaseTable):
    def __init__(
        self, headers=None, footers=None, column_types=None, name=None, settings=None,
    ):
        super().__init__(headers, footers, column_types, name, settings)
        self._setup(None)

    def _setup(self, obj):
        self._initialize()
